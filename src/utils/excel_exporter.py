import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def export_shiplogs_to_excel(logs, filepath):
    if not logs:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'No Data'
        ws['A1'] = 'Không có dữ liệu nhật ký phát hiện để xuất.'
        wb.save(filepath)
        return
    df = pd.DataFrame(logs)
    column_mapping = {'track_id': 'Mã Tracking ID', 'class_name': 'Loại Tàu', 'gio_phat_hien': 'Thời Gian Phát Hiện', 'video_source': 'Nguồn Video/Kênh', 'so_hieu_ocr': 'Số Hiệu Tàu (OCR)', 'do_tin_cay_ocr': 'Độ Tin Cậy OCR (%)', 'ghi_chu': 'Ghi Chú/Cảnh Báo'}
    cols_to_use = [c for c in column_mapping.keys() if c in df.columns]
    df_clean = df[cols_to_use].rename(columns={c: column_mapping[c] for c in cols_to_use})
    if 'Độ Tin Cậy OCR (%)' in df_clean.columns:
        df_clean['Độ Tin Cậy OCR (%)'] = df_clean['Độ Tin Cậy OCR (%)'].apply(lambda x: round(x * 100, 1) if pd.notnull(x) else 0.0)
    if 'Ghi Chú/Cảnh Báo' in df_clean.columns:
        df_clean['Ghi Chú/Cảnh Báo'] = df_clean['Ghi Chú/Cảnh Báo'].fillna('')
    wb = openpyxl.Workbook()
    ws_logs = wb.active
    ws_logs.title = 'Nhật Ký Chi Tiết'
    ws_logs.views.sheetView[0].showGridLines = True
    ws_logs.merge_cells('A1:G1')
    title_cell = ws_logs['A1']
    title_cell.value = 'BÁO CÁO NHẬT KÝ PHÁT HIỆN TÀU THUYỀN'
    title_cell.font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_logs.row_dimensions[1].height = 40
    headers = list(df_clean.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws_logs.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_logs.row_dimensions[3].height = 25
    thin_border = Border(left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'), top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'))
    for row_num, row_data in enumerate(df_clean.values, 4):
        ws_logs.row_dimensions[row_num].height = 20
        for col_num, val in enumerate(row_data, 1):
            cell = ws_logs.cell(row=row_num, column=col_num)
            cell.value = val
            cell.font = Font(name='Segoe UI', size=9)
            cell.border = thin_border
            col_name = headers[col_num - 1]
            if col_name in ['Mã Tracking ID', 'Độ Tin Cậy OCR (%)']:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_name in ['Loại Tàu', 'Thời Gian Phát Hiện', 'Ghi Chú/Cảnh Báo']:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            if col_name == 'Ghi Chú/Cảnh Báo' and val == '[CẢNH BÁO XÂM NHẬP]':
                cell.fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
                cell.font = Font(name='Segoe UI', size=9, bold=True, color='C0392B')
    ws_stats = wb.create_sheet(title='Thống Kê Tổng Quan')
    ws_stats.views.sheetView[0].showGridLines = True
    ws_stats.merge_cells('A1:D1')
    title_stats = ws_stats['A1']
    title_stats.value = 'THỐNG KÊ & PHÂN TÍCH TỔNG QUAN'
    title_stats.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
    title_stats.fill = PatternFill(start_color='16A085', end_color='16A085', fill_type='solid')
    title_stats.alignment = Alignment(horizontal='center', vertical='center')
    ws_stats.row_dimensions[1].height = 35
    total_detections = len(df)
    class_counts = df['class_name'].value_counts()
    violation_count = 0
    if 'ghi_chu' in df.columns:
        violation_count = df['ghi_chu'].fillna('').str.contains('CẢNH BÁO').sum()
    ws_stats['A3'] = 'Tổng số lượt phát hiện:'
    ws_stats['A3'].font = Font(name='Segoe UI', size=10, bold=True)
    ws_stats['B3'] = total_detections
    ws_stats['B3'].font = Font(name='Segoe UI', size=10, bold=True, color='2980B9')
    ws_stats['B3'].alignment = Alignment(horizontal='left')
    ws_stats['A4'] = 'Tổng số xâm nhập vùng cấm:'
    ws_stats['A4'].font = Font(name='Segoe UI', size=10, bold=True)
    ws_stats['B4'] = violation_count
    ws_stats['B4'].font = Font(name='Segoe UI', size=10, bold=True, color='C0392B')
    ws_stats['B4'].alignment = Alignment(horizontal='left')
    if violation_count > 0:
        ws_stats['B4'].fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    ws_stats['A6'] = 'Phân Loại Tàu'
    ws_stats['A6'].font = Font(name='Segoe UI', size=11, bold=True, color='16A085')
    ws_stats['A7'] = 'Loại Tàu'
    ws_stats['B7'] = 'Số Lượng'
    ws_stats['C7'] = 'Tỉ Lệ (%)'
    for col_letter in ['A7', 'B7', 'C7']:
        cell = ws_stats[col_letter]
        cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    ws_stats.row_dimensions[7].height = 22
    current_row = 8
    for cls_name, count in class_counts.items():
        ws_stats.row_dimensions[current_row].height = 18
        c1 = ws_stats.cell(row=current_row, column=1, value=cls_name)
        c2 = ws_stats.cell(row=current_row, column=2, value=count)
        pct = count / total_detections * 100 if total_detections > 0 else 0
        c3 = ws_stats.cell(row=current_row, column=3, value=f'{pct:.1f}%')
        for c in [c1, c2, c3]:
            c.font = Font(name='Segoe UI', size=9)
            c.border = thin_border
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c2.alignment = Alignment(horizontal='right', vertical='center')
        c3.alignment = Alignment(horizontal='right', vertical='center')
        current_row += 1
    for ws in [ws_logs, ws_stats]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1 or cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
    wb.save(filepath)
    print(f'>> Excel Report exported successfully to {filepath}')
